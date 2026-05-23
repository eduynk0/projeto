from flask import Flask, render_template, request, redirect, send_file
import os
import csv
from datetime import datetime
import openpyxl

# ======================================
# CONFIGURAÇÃO INICIAL
# ======================================

app = Flask(__name__)

PASTA_CLIENTES = "clientes"

# cria pasta clientes caso não exista
if not os.path.exists(PASTA_CLIENTES):
    os.makedirs(PASTA_CLIENTES)

def get_proximo_pedido():
    total = 0
    if os.path.exists(PASTA_CLIENTES):
        for root, dirs, files in os.walk(PASTA_CLIENTES):
            total += len([f for f in files if f.endswith('.csv')])
    return total + 2000

# ======================================
# PÁGINA INICIAL
# ======================================

@app.route("/")
def index():
    # Lista todas as subpastas dentro da pasta 'clientes'
    clientes_existentes = []
    historico_pedidos = []
    
    if os.path.exists(PASTA_CLIENTES):
        clientes_existentes = [d for d in os.listdir(PASTA_CLIENTES) if os.path.isdir(os.path.join(PASTA_CLIENTES, d))]
        
        # Busca histórico de pedidos percorrendo as pastas
        for cliente in clientes_existentes:
            caminho_p = os.path.join(PASTA_CLIENTES, cliente)
            arquivos = [f for f in os.listdir(caminho_p) if f.endswith('.csv')]
            for arq in arquivos:
                caminho_arq = os.path.join(caminho_p, arq)
                data_f = "---"
                with open(caminho_arq, newline="", encoding="utf-8-sig") as f:
                    leitor = csv.DictReader(f)
                    linha = next(leitor, {})
                    data_f = linha.get("data_hora", "---")

                # Extrai o número do pedido do nome do arquivo (ex: pedido_2000.csv)
                num = arq.replace("pedido_", "").replace(".csv", "")
                historico_pedidos.append({
                    "numero": num,
                    "ano": data_f.split('/')[2].split(' ')[0] if data_f != "---" else datetime.now().year,
                    "cliente_slug": cliente,
                    "cliente_nome": cliente.capitalize(),
                    "arquivo": arq,
                    "data": data_f
                })
    
    # Ordena o histórico pelo número do pedido (mais recente primeiro)
    historico_pedidos.sort(key=lambda x: int(x['numero']), reverse=True)
    
    # Limita a exibição aos 20 pedidos mais recentes para manter a performance
    historico_pedidos = historico_pedidos[:20]

    proximo = get_proximo_pedido()
    return render_template("index.html", clientes=clientes_existentes, proximo_pedido=proximo, historico=historico_pedidos, ano_atual=datetime.now().year)

# ======================================
# BUSCAR OU CRIAR CLIENTE
# ======================================

@app.route("/buscar", methods=["POST"])
def buscar_cliente():

    nome_cliente = request.form["cliente"]

    nome_cliente = nome_cliente.strip().lower()

    pasta_cliente = os.path.join(
        PASTA_CLIENTES,
        nome_cliente
    )

    # cria pasta se não existir
    if not os.path.exists(pasta_cliente):

        os.makedirs(pasta_cliente)

    return redirect(f"/cadastro/{nome_cliente}")

# ======================================
# PÁGINA DE CADASTRO
# ======================================

@app.route("/cadastro/<cliente>")
def cadastro(cliente):

    dados = {
        "cliente": cliente,
        "telefone": "",
        "endereco": "",
        "cidade": "",
        "responsavel": "",
        "email": "",
        "data_hora": ""
    }

    pasta_cliente = os.path.join(
        PASTA_CLIENTES,
        cliente
    )

    arquivos = os.listdir(pasta_cliente)

    arquivos_csv = []

    for arquivo in arquivos:

        if arquivo.endswith(".csv"):

            arquivos_csv.append(arquivo)

    # verifica se já existem pedidos
    if len(arquivos_csv) > 0:

        ultimo_arquivo = sorted(arquivos_csv)[-1]

        caminho_arquivo = os.path.join(
            pasta_cliente,
            ultimo_arquivo
        )

        with open(
            caminho_arquivo,
            newline="",
            encoding="utf-8-sig"
        ) as arquivo:

            leitor = csv.DictReader(arquivo)

            for linha in leitor:
                # Pega dados de contato do último arquivo, mas não os produtos
                dados = linha.copy()
                # Limpa os campos específicos do pedido anterior para o novo orçamento
                dados["valor_total_pedido"] = "0.00"
                dados["data_hora"] = ""
                break
    
    proximo = get_proximo_pedido()

    return render_template(
        "cadastro.html",
        dados=dados,
        proximo_pedido=proximo,
        produtos=[],
        ano_atual=datetime.now().year
    )

@app.route("/editar/<cliente>/<arquivo_nome>")
def editar(cliente, arquivo_nome):
    pasta_cliente = os.path.join(PASTA_CLIENTES, cliente)
    caminho = os.path.join(pasta_cliente, arquivo_nome)
    
    produtos = []
    dados = {}
    
    if os.path.exists(caminho):
        with open(caminho, newline="", encoding="utf-8-sig") as f:
            leitor = csv.DictReader(f)
            for linha in leitor:
                if not dados:
                    dados = linha
                produtos.append(linha)
    
    num_pedido = arquivo_nome.replace("pedido_", "").replace(".csv", "")
    ano_pedido = dados.get("data_hora", "").split('/')[2].split(' ')[0] if dados.get("data_hora") else datetime.now().year
    
    return render_template(
        "cadastro.html", 
        dados=dados, 
        produtos=produtos, 
        proximo_pedido=num_pedido, 
        nome_arquivo=arquivo_nome,
        ano_atual=ano_pedido
    )

# ======================================
# EXPORTAR PARA EXCEL (XLSX)
# ======================================

@app.route("/exportar/<cliente>/<arquivo_nome>")
def exportar(cliente, arquivo_nome):
    pasta_cliente = os.path.join(PASTA_CLIENTES, cliente)
    caminho_csv = os.path.join(pasta_cliente, arquivo_nome)
    caminho_modelo = "modelo.xlsx" # O arquivo deve estar na pasta raiz do projeto

    if not os.path.exists(caminho_modelo):
        return "Erro: Arquivo 'modelo.xlsx' não encontrado na raiz do projeto.", 404

    # 1. Ler dados do CSV
    produtos = []
    dados_gerais = {}
    if os.path.exists(caminho_csv):
        with open(caminho_csv, newline="", encoding="utf-8-sig") as f:
            leitor = csv.DictReader(f)
            for linha in leitor:
                if not dados_gerais:
                    dados_gerais = linha
                produtos.append(linha)

    # 2. Abrir o Excel e preencher
    wb = openpyxl.load_workbook(caminho_modelo)
    ws = wb.active

    # Calcula a ordem do pedido específica para este cliente (1, 2, 3...)
    arquivos_cliente = sorted([f for f in os.listdir(pasta_cliente) if f.endswith('.csv')])
    try:
        ordem_cliente = arquivos_cliente.index(arquivo_nome) + 1
    except ValueError:
        ordem_cliente = 1

    ano_pedido = dados_gerais.get("data_hora", "").split('/')[2].split(' ')[0] if dados_gerais.get("data_hora") else datetime.now().year

    # Mapa de substituição para campos simples
    mapa = {
        "{cliente}": dados_gerais.get("cliente", "").capitalize(),
        "{telefone}": dados_gerais.get("telefone", ""),
        "{endereco}": dados_gerais.get("endereco", ""),
        "{cidade}": dados_gerais.get("cidade", ""),
        "{responsavel}": dados_gerais.get("responsavel", ""),
        "{email}": dados_gerais.get("email", ""),
        "{valor_total_pedido}": dados_gerais.get("valor_total_pedido", "0.00"),
        "{data_hora}": dados_gerais.get("data_hora", ""),
        "{pedido_numero}": ordem_cliente,
        "{ano}": ano_pedido
    }

    # Varre as células para substituições simples
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                for chave, valor in mapa.items():
                    if chave in cell.value:
                        cell.value = cell.value.replace(chave, str(valor))

    # Lógica para preencher a tabela de produtos
    # Procuramos a linha que contém o placeholder {produto}
    linha_tabela = -1
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "{produto}":
                linha_tabela = cell.row
                break
        if linha_tabela != -1: break

    if linha_tabela != -1:
        for i, prod in enumerate(produtos):
            row_atual = linha_tabela + i
            if i > 0: ws.insert_rows(row_atual) # Insere nova linha se houver mais de um item
            
            ws.cell(row=row_atual, column=1).value = prod['produto']
            ws.cell(row=row_atual, column=2).value = float(prod['quantidade'])
            ws.cell(row=row_atual, column=3).value = float(prod['valor_unitario'])

    # 3. Salvar e enviar para download
    nome_saida = f"{cliente}({ordem_cliente}).xlsx"
    caminho_saida = os.path.join(pasta_cliente, nome_saida)
    wb.save(caminho_saida)

    return send_file(caminho_saida, as_attachment=True)

# ======================================
# SALVAR PEDIDO
# ======================================

@app.route("/salvar", methods=["POST"])
def salvar():

    cliente = request.form["cliente"]

    cliente = cliente.strip().lower()

    pasta_cliente = os.path.join(
        PASTA_CLIENTES,
        cliente
    )

    # Se vier um nome de arquivo, estamos editando. Caso contrário, criamos um novo.
    nome_arquivo = request.form.get("nome_arquivo")
    if not nome_arquivo:
        numero_pedido = get_proximo_pedido()
        nome_arquivo = f"pedido_{numero_pedido:03}.csv"

    caminho_arquivo = os.path.join(
        pasta_cliente,
        nome_arquivo
    )
    
    # Captura as listas de produtos do formulário
    produtos = request.form.getlist("produto[]")
    quantidades = request.form.getlist("quantidade[]")
    valores = request.form.getlist("valor[]")
    valor_total_geral = request.form.get("valor_total_geral")
    data_atual = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    campos = [
        "cliente",
        "telefone",
        "endereco",
        "cidade",
        "responsavel",
        "email",
        "produto",
        "quantidade",
        "valor_unitario",
        "valor_total_pedido",
        "data_hora"
    ]

    with open(
        caminho_arquivo,
        "w",
        newline="",
        encoding="utf-8-sig"
    ) as arquivo:

        writer = csv.DictWriter(
            arquivo,
            fieldnames=campos
        )

        writer.writeheader()

        # Salva uma linha para cada produto, mantendo os dados do cliente e o total geral
        for i in range(len(produtos)):
            writer.writerow({
                "cliente": request.form["cliente"],
                "telefone": request.form["telefone"],
                "endereco": request.form["endereco"],
                "cidade": request.form["cidade"],
                "responsavel": request.form["responsavel"],
                "email": request.form["email"],
                "produto": produtos[i],
                "quantidade": quantidades[i],
                "valor_unitario": valores[i],
                "valor_total_pedido": valor_total_geral,
                "data_hora": data_atual
            })

    # Se o usuário clicou em 'Salvar e Exportar', redireciona para a rota de exportação
    if request.form.get("acao") == "exportar":
        return redirect(f"/exportar/{cliente}/{nome_arquivo}")

    return redirect("/")

# ======================================
# EXECUTAR SERVIDOR
# ======================================

if __name__ == "__main__":

    app.run(debug=True)